from fileinput import filename
import random
from sre_constants import SUCCESS
import time
from tkinter.filedialog import SaveAs
import numpy as np
from pandas import pandas as pd, DataFrame 
from openpyxl.styles import PatternFill,Font, Alignment,Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, Rule
from openpyxl import load_workbook
from sys import executable
import PySimpleGUI as sg
import os




# srcFile  = 'inputFiles\source1.xlsx'
# srcSheet = 'source' 
# srcId = 'GDMFSL_EXTERNAL_ID__C'
 
# trgtFile = 'inputFiles\\target1.xlsx'
# trgtSheet = 'target1'     
# trgtId = 'GDMFSL_External_ID__c'

# exportfileName = 'outputFiles\merged.xlsx'


def import_trgt_to_src(src, trgt, srcId, trgtId, exportfileName):
   
    print("Merging source and target sheets according to Primary key")    
    mergedFile = src.merge(trgt, left_on = srcId.upper(), right_on = trgtId.lower(), how="left" )
    # print(mergedFile)
    print('Merging completed \nExporting as excel file\n\n',mergedFile)

    mergedFile.to_excel(exportfileName)
    print("done!")
    return mergedFile

def validate2(src, trgt, trgtId, mergedFile, exportfileName):

    print("Validating...")
    srcCols = src.columns
    trgtCols = trgt.columns
    src2 =  mergedFile[srcCols]
    trgt2 =  mergedFile[trgtCols]
   
    columns = src2.columns = trgt2.columns
    resultSet = src2 == trgt2
    print(resultSet)
    mismatchCount = dict()

    print("finding mismatchs...")
    srcCount = len(resultSet[trgtId.lower()])
    trgtCount = (resultSet[trgtId.lower()]==True).sum()
    print("Src=",srcCount, "trgt=",trgtCount)
    # print(type(columns))
    columns = list(columns)
    columns.remove(trgtId.lower())
    filt = (resultSet[trgtId.lower()]==True)
    resultSet2 = resultSet[filt]
    for i in columns:
        if (resultSet2[i]==False).sum()>0:          
            mismatchRowNum1 = resultSet2.index[resultSet2[i] == False].tolist()
            # print(resultSet[i] ,"\n",mismatchRowNum1)
            mismatchRowNum = random.choice(mismatchRowNum1)
            # print(mismatchRowNum)
            srcMismatchValue = src2.at[mismatchRowNum,i]
            trgtMismatchValue = trgt2.at[mismatchRowNum,i]
            count=(resultSet2[i]==False).sum()
            mismatchCount[i] = f"Mismatchs: {count}"+f"\nEg at row {mismatchRowNum+2}:"+f"\nSrc={srcMismatchValue}"+f"\nTrgt={trgtMismatchValue}"
            
    output = pd.concat([mergedFile, resultSet], axis = 1)
    output.to_excel(exportfileName)
    print("Done!")
    return mismatchCount, output, srcCount, trgtCount
    

def styling(outputFileName):
    print('Adding colors and styling')
    final_wb = load_workbook(outputFileName)
    final_ws = final_wb.active
    # final_ws.delete_cols(1,1)

    cols = round(final_ws.max_column/3)
    #fill yellow colors to src
    for col in range(1,cols+1):
        final_ws.cell(1, col).fill = PatternFill(bgColor="00FFFF00", fill_type = "gray125")
        final_ws.cell(1, col).font = Font(size=12, bold=True)

    #filling target colors
    for col in range(1,cols+1):
        final_ws.cell(1, col+cols).fill = PatternFill(bgColor="00008080", fill_type = "gray125")
        final_ws.cell(1, col+cols).font = Font(size=12, bold=True)
    #filling validation colors
    for col in range(1,cols+1):
        final_ws.cell(1, col+cols*2).fill = PatternFill(bgColor="0000FF00", fill_type = "gray125")
        final_ws.cell(1, col+cols*2).font = Font(size=12, bold=True)

    for i in range(2, final_ws.max_row+1):
        for j in range(cols*2+1, final_ws.max_column+1):
            # temp = final_ws.cell(i,j).value 
            # print(temp)
            if (final_ws.cell(i,j).value == True):
                final_ws.cell(i, j).font = Font(color="00008000")
            else: 
                final_ws.cell(i, j).font = Font(color="00FF0000")

    final_ws.insert_cols(cols+1)
    final_ws.insert_cols(cols*2+2) 
    
    obs_ws = final_wb["Observations"]
    all = Side(border_style='thin')
    border = Border(left=all, right=all, top=all, bottom=all)
    for col in obs_ws.iter_cols():        
        for cell in col:
            cell.alignment=Alignment( wrapText = True )
            cell.border=border

    obs_ws.column_dimensions['A'].width = 4
    obs_ws.column_dimensions['B'].width = 25
    obs_ws.column_dimensions['C'].width = 25

    print(f"Saving file as {outputFileName}\nDone!")
    final_wb.save(outputFileName)
    sg.popup('success','success')
    

def take_observations(exportfileName, srcCount, trgtCount, mismatchCount, validaitonOutput):

    column = ["S.No","Field_Name", "Comments"]
    totals = f"Source={srcCount} \nTarget={trgtCount}"
    data = {
            "S.No" : [i for i in range(1,len(mismatchCount)+1)],
            "Field_Name" : [ keys.upper() for keys in mismatchCount],
            "Comments"  :[ values for values in mismatchCount.values()]
        }
    # print(data)
    data["S.No"].insert(0, "")
    data["Field_Name"].insert(0, totals)
    data["Comments"].insert(0, "")
    df = DataFrame(columns=column, index=[i for i in range(len(mismatchCount)+1)])
    for i in data:
        df[i] = data[i]

    with pd.ExcelWriter(exportfileName, engine='xlsxwriter') as writer:
        validaitonOutput.to_excel(writer, sheet_name="Validation", index=False)
        df.to_excel(writer, sheet_name="Observations", index=False)

    
# main Functions
def submit(srcFile, trgtFile, srcSheet, trgtSheet, srcId, trgtId,exportfileName):
    src = pd.read_excel(srcFile, srcSheet)
    src.columns = src.columns.str.upper()
    src.sort_index(axis = 1)
    print(src)

    trgt = pd.read_excel(trgtFile, trgtSheet)
    trgt.columns = trgt.columns.str.lower()
    trgt.sort_index(axis = 1)
    print(trgt)

    # importing data
    importedData  = import_trgt_to_src(src, trgt, srcId, trgtId, exportfileName)
    mismatchCount, validationOutput, srcCount, trgtCount = validate2(src, trgt,trgtId, importedData, exportfileName)
    take_observations(exportfileName,srcCount, trgtCount, mismatchCount, validationOutput )
    styling(exportfileName)

    # printing details
    print('total (rows,cols) in src:', src.shape)
    # print('total rows in src:', len(src.index))
    print('total (rows, cols) in trgt:', trgt.shape)
    # print('total rows in trgt:', len(trgt.index))
    print('total (rows, cols) in merge:', importedData.shape)
    # print('total rows in merge:', len(importedData.index))
sg.theme('LightBlue7')
sg.Titlebar('VALIDATION')
working_dir = os.getcwd()

src_sheet_names = []
trgt_sheet_names = []

layout =[
    # Choose Files
    [sg.Text('Choose Source file: ', size=(15,1)), sg.InputText(key="srcFile" ), sg.FileBrowse(initial_folder=working_dir)],
    [sg.Text('Choose Target file: ', size=(15,1)), sg.InputText(key="trgtFile"), sg.FileBrowse(initial_folder=working_dir)],
    # List boxes for id and sheet name
    [sg.Text('Source Sheet:', size=(15, 1)), sg.InputText( key='srcSheet', size=(15, 1))],
    [sg.Text('Target Sheet:', size=(15, 1)), sg.InputText( key='trgtSheet', size=(15, 1))],

    [sg.Text('Source Primary key: '), sg.InputText(key='srcId', size=(15, 1))],
    [sg.Text('Target Primary key: '), sg.InputText( key='trgtId', size=(15, 1))],

    [sg.Text('Save output as: ', size=(15, 1)), sg.InputText(key='exportfileName')],
    [sg.Submit("Validate"), sg.Exit()]]
    
    # calculating process time

def main():
    window = sg.Window('file validation',layout)
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED or event == 'Exit':
            break
        
        elif event == 'Validate':
            srcFile = values['srcFile']
            trgtFile = values['trgtFile']
            srcSheet = values['srcSheet']
            trgtSheet = values['trgtSheet']
            srcId = values['srcId']
            trgtId = values['trgtId']
            exportfilename = values['exportfileName']
            if srcFile and trgtFile and srcSheet and trgtSheet and srcId and trgtId and exportfilename:
                submit(srcFile, trgtFile, srcSheet, trgtSheet, srcId, trgtId, exportfilename)
            else: 
                sg.popup('enter all fields')
        print()
        print(event, values)

    window.close()


if __name__ == '__main__':
    main()





"""
# srcFile  = 'xlFiles\source2.xlsx'
# srcSheet = 'source' #ACCOUNTID
# srcId = 'ACCOUNTID'

# trgtFile = 'xlFiles\Target2 .xlsx'
# trgtSheet = 'bulkQuery_result_7505a00000epjV'   #ACCOUNTID
# exportfileName = 'merged.xlsx'
# trgtId = 'Accountid'

# f = open('Report1.txt', 'a')
 # f.write(str(temp1)+" \t " +str(temp2)+"\n")
#    f.close()




# srcFile  = 'xlFiles\source5.xlsx'
# srcSheet = 'Sheet1' #ACCOUNTID
# srcId = 'GDMFSL_OBJECT_NUMBER__C'

# trgtFile = 'xlFiles\\target5.xlsx'
# trgtSheet = 'Sheet1'   #ACCOUNTID
# exportfileName = 'merged.xlsx'
# trgtId = 'GDMFSL_OBJECT_NUMBER__C'

# srcFile  = 'xlFiles\source2.xlsx'
# srcSheet = 'source' #ACCOUNTID
# srcId = 'ACCOUNTID'

# trgtFile = 'xlFiles\Target2 .xlsx'
# trgtSheet = 'bulkQuery_result_7505a00000epjV'   #ACCOUNTID
# exportfileName = 'merged.xlsx'
# trgtId = 'Accountid'




"""


